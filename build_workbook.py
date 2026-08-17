"""Export the campaign results as one .xlsx workbook.

build_results.py writes one CSV pair per experiment; this collapses them into a
single file with a tab per arm, plus a Summary tab that puts every controller's
clean / C2 / C3 numbers in one table.

Tabs:
  Summary              one row per controller x arm, with the ATT delta and how
                       many clean seed spreads it is worth
  <arm> by config      seeds collapsed to mean/std (build_results' config sheet)
  <arm> runs           one row per run (the physical inventory)

Numbers come from the run manifests via build_results, not from the CSVs, so
the workbook can never be stale relative to logs/.

    python build_workbook.py
    python build_workbook.py --experiments c2_campaign c3_campaign
    python build_workbook.py --out report_sheet/blockage_campaign.xlsx
"""
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import build_results as br
from configurations import LOGS_DIR_NAME
from experiments import CONTROLLERS

DEFAULT_EXPERIMENTS = ["c2_campaign", "c3_campaign"]
DEFAULT_OUT = "report_sheet/blockage_campaign.xlsx"

# A blockage arm is accepted when its ATT delta is worth at least this many
# clean seed spreads -- the rule the c3_decision gate was built around.
ACCEPTANCE_SPREAD_MULTIPLE = 5

HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
PASS_FILL = PatternFill("solid", fgColor="D9EAD3")

SUMMARY_COLUMNS = [
    ("controller", "controller"),
    ("arm", "arm"),
    ("scenario", "scenario"),
    ("seeds", "n_seeds"),
    ("ATT (s)", "att_mean"),
    ("ATT sd", "att_std"),
    ("AQL", "aql_mean"),
    ("completion", "completion_mean"),
    ("delta ATT (s)", "delta_att"),
    ("delta / clean spread", "delta_over_spread"),
]


def number_format(column):
    if column in ("completion", "completion_mean", "completion_rate"):
        return "0.000"
    if any(k in column for k in ("att", "aql", "delta", "awt", "spread")):
        return "0.0"
    if any(k in column for k in ("finished", "not_inserted", "teleports")):
        return "0"
    return None


def load_experiment(experiment, logs_dir):
    """(run rows, per-config rows) for one experiment, straight from manifests."""
    scope_dir = Path(logs_dir) / experiment
    if not scope_dir.exists():
        return None, None
    rows = br.collect_rows(scope_dir, Path(logs_dir))
    configs = br.aggregate_configs(br.dedupe_completed(rows))
    return rows, configs


def arm_of(experiment):
    return experiment.split("_")[0]


def controller_order(controller):
    names = list(CONTROLLERS)
    return names.index(controller) if controller in names else len(names)


def summary_row(config, arm):
    return {
        "controller": config["controller"],
        "arm": arm,
        "scenario": config["blockage"],
        "n_seeds": config["n_seeds"],
        "att_mean": config["att_internal_mean"],
        "att_std": config["att_internal_std"],
        "aql_mean": config["aql_mean"],
        "completion_mean": config["completion_rate_mean"],
        "delta_att": config["delta_att_internal"],
        "delta_over_spread": config["delta_over_clean_spread"],
    }


def build_summary(configs_by_experiment):
    """One clean row per controller, then one row per controller x arm.

    The clean arm is identical across experiments (same controller, seeds and
    routes, no blockage), so it is emitted once -- from the first experiment
    that has it, with a warning if a later one disagrees.
    """
    clean_by_controller = {}
    blockage_rows = []
    for experiment, configs in configs_by_experiment.items():
        arm = arm_of(experiment)
        for config in configs:
            if config["blockage"] == "none":
                seen = clean_by_controller.get(config["controller"])
                if seen is None:
                    clean_by_controller[config["controller"]] = summary_row(
                        config, "clean")
                elif _differs(seen["att_mean"], config["att_internal_mean"]):
                    print(f"  warning: {config['controller']} clean ATT differs "
                          f"between experiments ({seen['att_mean']} vs "
                          f"{config['att_internal_mean']}); keeping the first.")
            else:
                blockage_rows.append(summary_row(config, arm))

    rows = []
    for controller in sorted(clean_by_controller, key=controller_order):
        rows.append(clean_by_controller[controller])
        rows += sorted((r for r in blockage_rows if r["controller"] == controller),
                       key=lambda r: r["arm"])
    return rows


def _differs(a, b, tolerance=0.05):
    try:
        return abs(float(a) - float(b)) > tolerance
    except (TypeError, ValueError):
        return a != b


def write_sheet(workbook, title, headers, keys, rows, highlight_key=None):
    sheet = workbook.create_sheet(title[:31])
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(key, "") for key in keys])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    sheet.freeze_panes = "A2"

    for index, key in enumerate(keys, start=1):
        fmt = number_format(key)
        if fmt:
            for cell in sheet[get_column_letter(index)][1:]:
                cell.number_format = fmt
    if highlight_key in keys:
        _highlight_accepted(sheet, keys.index(highlight_key) + 1)
    _autosize(sheet, headers)
    return sheet


def _highlight_accepted(sheet, column_index):
    """Green the arms whose delta clears the acceptance rule."""
    for cell in sheet[get_column_letter(column_index)][1:]:
        if isinstance(cell.value, (int, float)) and \
                cell.value >= ACCEPTANCE_SPREAD_MULTIPLE:
            cell.fill = PASS_FILL
            cell.font = Font(bold=True)


def _autosize(sheet, headers, cap=42):
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        widest = max([len(str(header))]
                     + [len(str(c.value)) for c in sheet[letter][1:]
                        if c.value is not None] or [0])
        sheet.column_dimensions[letter].width = min(widest + 2, cap)


def build_workbook(experiments, logs_dir):
    workbook = Workbook()
    workbook.remove(workbook.active)

    configs_by_experiment = {}
    for experiment in experiments:
        rows, configs = load_experiment(experiment, logs_dir)
        if rows is None:
            print(f"  skipping {experiment}: no logs/{experiment}/ directory.")
            continue
        configs_by_experiment[experiment] = configs
        print(f"  {experiment}: {len(rows)} runs, {len(configs)} configs")

    if not configs_by_experiment:
        return None

    summary = build_summary(configs_by_experiment)
    headers = [h for h, _ in SUMMARY_COLUMNS]
    keys = [k for _, k in SUMMARY_COLUMNS]
    write_sheet(workbook, "Summary", headers, keys, summary,
                highlight_key="delta_over_spread")

    for experiment, configs in configs_by_experiment.items():
        arm = arm_of(experiment)
        write_sheet(workbook, f"{arm} by config", br.CONFIG_COLUMNS,
                    br.CONFIG_COLUMNS, configs,
                    highlight_key="delta_over_clean_spread")
        runs, _ = load_experiment(experiment, logs_dir)
        write_sheet(workbook, f"{arm} runs", br.RUN_COLUMNS, br.RUN_COLUMNS, runs)
    return workbook


def parse_args():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--experiments", nargs="+", default=DEFAULT_EXPERIMENTS,
                        help=f"Experiments to include (default: "
                             f"{DEFAULT_EXPERIMENTS}).")
    parser.add_argument("--out", default=DEFAULT_OUT,
                        help=f"Workbook path (default: {DEFAULT_OUT}).")
    parser.add_argument("--logs_dir", default=LOGS_DIR_NAME)
    return parser.parse_args()


def main(args):
    workbook = build_workbook(args.experiments, args.logs_dir)
    if workbook is None:
        raise SystemExit("Nothing to export: none of those experiments have logs.")
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)
    print(f"\nWrote {out} ({len(workbook.sheetnames)} tabs: "
          f"{', '.join(workbook.sheetnames)})")


if __name__ == "__main__":
    main(parse_args())
